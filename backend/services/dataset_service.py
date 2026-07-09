import logging
import io
from typing import Tuple
from sqlalchemy.orm import Session
from models.customer import Customer
from models.policy import Policy
from models.claim import Claim
from models.renewal import Renewal
from services.auth_service import hash_password

logger = logging.getLogger(__name__)

DATASET_SCHEMAS = {
    "customers": {"required": ["email", "full_name"], "optional": ["phone", "address", "age", "occupation", "risk_profile", "date_of_birth"]},
    "policies": {"required": ["policy_number", "customer_id", "policy_type", "coverage_amount", "premium_amount", "start_date", "end_date"], "optional": ["deductible", "status", "description", "beneficiary"]},
    "claims": {"required": ["claim_number", "customer_id", "policy_id"], "optional": ["claim_type", "amount", "status", "description", "filed_date"]},
    "renewals": {"required": ["policy_id", "customer_id"], "optional": ["renewal_date", "new_premium", "new_end_date", "status", "recommendation_score"]},
}


def _read_file(content: bytes, filename: str):
    ext = filename.rsplit(".", 1)[-1].lower()
    if ext == "csv":
        import csv
        text = content.decode("utf-8", errors="ignore")
        reader = csv.DictReader(io.StringIO(text))
        return list(reader)
    elif ext in ("xlsx", "xls"):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content))
            ws = wb.active
            headers = [str(cell.value) for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                rows.append(dict(zip(headers, [str(v) if v is not None else "" for v in row])))
            return rows
        except ImportError:
            raise ValueError("openpyxl is required for XLSX support. Install it via pip.")
    else:
        raise ValueError(f"Unsupported file type: {ext}")


def _detect_dataset_type(filename: str, rows: list) -> str:
    name = filename.lower().replace(".csv", "").replace(".xlsx", "")
    for key in DATASET_SCHEMAS:
        if key in name:
            return key
    if rows:
        cols = set(rows[0].keys())
        for key, schema in DATASET_SCHEMAS.items():
            if all(r in cols for r in schema["required"]):
                return key
    return ""


def _safe(row: dict, key: str, default=None):
    val = row.get(key, default)
    if val == "" or val is None:
        return default
    return val


def import_customers(rows: list, db: Session) -> Tuple[int, int]:
    imported, failed = 0, 0
    for row in rows:
        try:
            email = _safe(row, "email")
            if not email:
                failed += 1
                continue
            existing = db.query(Customer).filter(Customer.email == email).first()
            if existing:
                existing.full_name = _safe(row, "full_name", existing.full_name)
                existing.phone = _safe(row, "phone", existing.phone)
                existing.address = _safe(row, "address", existing.address)
                existing.age = int(_safe(row, "age", 0)) or existing.age
                existing.occupation = _safe(row, "occupation", existing.occupation)
                existing.risk_profile = _safe(row, "risk_profile", existing.risk_profile)
                existing.date_of_birth = _safe(row, "date_of_birth", existing.date_of_birth)
            else:
                db.add(Customer(
                    email=email,
                    full_name=_safe(row, "full_name", "Unknown"),
                    hashed_password=hash_password("Welcome@123"),
                    phone=_safe(row, "phone"),
                    address=_safe(row, "address"),
                    age=int(_safe(row, "age", 0)) or None,
                    occupation=_safe(row, "occupation"),
                    risk_profile=_safe(row, "risk_profile", "medium"),
                    date_of_birth=_safe(row, "date_of_birth"),
                ))
            imported += 1
        except Exception as e:
            logger.warning(f"Customer row failed: {e}")
            failed += 1
    db.commit()
    return imported, failed


def import_policies(rows: list, db: Session) -> Tuple[int, int]:
    imported, failed = 0, 0
    for row in rows:
        try:
            pnum = _safe(row, "policy_number")
            if not pnum:
                failed += 1
                continue
            existing = db.query(Policy).filter(Policy.policy_number == pnum).first()
            if existing:
                existing.customer_id = int(_safe(row, "customer_id", existing.customer_id))
                existing.policy_type = _safe(row, "policy_type", existing.policy_type)
                existing.coverage_amount = float(_safe(row, "coverage_amount", existing.coverage_amount))
                existing.premium_amount = float(_safe(row, "premium_amount", existing.premium_amount))
                existing.deductible = float(_safe(row, "deductible", existing.deductible))
                existing.start_date = _safe(row, "start_date", existing.start_date)
                existing.end_date = _safe(row, "end_date", existing.end_date)
                existing.status = _safe(row, "status", existing.status)
                existing.description = _safe(row, "description", existing.description)
                existing.beneficiary = _safe(row, "beneficiary", existing.beneficiary)
            else:
                db.add(Policy(
                    policy_number=pnum,
                    customer_id=int(_safe(row, "customer_id", 0)),
                    policy_type=_safe(row, "policy_type", "health"),
                    coverage_amount=float(_safe(row, "coverage_amount", 0)),
                    premium_amount=float(_safe(row, "premium_amount", 0)),
                    deductible=float(_safe(row, "deductible", 0)),
                    start_date=_safe(row, "start_date", ""),
                    end_date=_safe(row, "end_date", ""),
                    status=_safe(row, "status", "active"),
                    description=_safe(row, "description"),
                    beneficiary=_safe(row, "beneficiary"),
                ))
            imported += 1
        except Exception as e:
            logger.warning(f"Policy row failed: {e}")
            failed += 1
    db.commit()
    return imported, failed


def import_claims(rows: list, db: Session) -> Tuple[int, int]:
    imported, failed = 0, 0
    for row in rows:
        try:
            cnum = _safe(row, "claim_number")
            if not cnum:
                failed += 1
                continue
            existing = db.query(Claim).filter(Claim.claim_number == cnum).first()
            if existing:
                existing.customer_id = int(_safe(row, "customer_id", existing.customer_id))
                existing.policy_id = int(_safe(row, "policy_id", existing.policy_id))
                existing.claim_type = _safe(row, "claim_type", existing.claim_type)
                existing.amount = float(_safe(row, "amount", 0)) or existing.amount
                existing.status = _safe(row, "status", existing.status)
                existing.description = _safe(row, "description", existing.description)
                existing.filed_date = _safe(row, "filed_date", existing.filed_date)
            else:
                db.add(Claim(
                    claim_number=cnum,
                    customer_id=int(_safe(row, "customer_id", 0)),
                    policy_id=int(_safe(row, "policy_id", 0)),
                    claim_type=_safe(row, "claim_type"),
                    amount=float(_safe(row, "amount", 0)) or None,
                    status=_safe(row, "status", "pending"),
                    description=_safe(row, "description"),
                    filed_date=_safe(row, "filed_date"),
                ))
            imported += 1
        except Exception as e:
            logger.warning(f"Claim row failed: {e}")
            failed += 1
    db.commit()
    return imported, failed


def import_renewals(rows: list, db: Session) -> Tuple[int, int]:
    imported, failed = 0, 0
    for row in rows:
        try:
            policy_id = int(_safe(row, "policy_id", 0))
            customer_id = int(_safe(row, "customer_id", 0))
            existing = db.query(Renewal).filter(
                Renewal.policy_id == policy_id,
                Renewal.customer_id == customer_id
            ).first()
            if existing:
                existing.renewal_date = _safe(row, "renewal_date", existing.renewal_date)
                existing.new_premium = float(_safe(row, "new_premium", 0)) or existing.new_premium
                existing.new_end_date = _safe(row, "new_end_date", existing.new_end_date)
                existing.status = _safe(row, "status", existing.status)
                existing.recommendation_score = float(_safe(row, "recommendation_score", 0)) or existing.recommendation_score
            else:
                db.add(Renewal(
                    policy_id=policy_id,
                    customer_id=customer_id,
                    renewal_date=_safe(row, "renewal_date"),
                    new_premium=float(_safe(row, "new_premium", 0)) or None,
                    new_end_date=_safe(row, "new_end_date"),
                    status=_safe(row, "status", "pending"),
                    recommendation_score=float(_safe(row, "recommendation_score", 0)) or None,
                ))
            imported += 1
        except Exception as e:
            logger.warning(f"Renewal row failed: {e}")
            failed += 1
    db.commit()
    return imported, failed


def import_dataset(content: bytes, filename: str, db: Session) -> dict:
    rows = _read_file(content, filename)
    if not rows:
        return {"status": "failed", "message": "File is empty or unreadable.", "imported": 0, "failed": 0}

    dataset_type = _detect_dataset_type(filename, rows)
    if not dataset_type:
        return {"status": "failed", "message": "Invalid dataset structure. Could not detect dataset type.", "imported": 0, "failed": 0}

    schema = DATASET_SCHEMAS[dataset_type]
    cols = set(rows[0].keys())
    missing = [r for r in schema["required"] if r not in cols]
    if missing:
        return {"status": "failed", "message": f"Invalid dataset structure. Missing columns: {', '.join(missing)}", "imported": 0, "failed": 0}

    importers = {
        "customers": import_customers,
        "policies": import_policies,
        "claims": import_claims,
        "renewals": import_renewals,
    }
    imported, failed = importers[dataset_type](rows, db)
    return {
        "status": "imported",
        "dataset_type": dataset_type,
        "filename": filename,
        "total_rows": len(rows),
        "imported": imported,
        "failed": failed,
        "message": f"{imported} records upserted (inserted or updated), {failed} failed.",
    }


def get_import_statistics(db: Session) -> dict:
    from models.notification import Notification
    return {
        "customers": db.query(Customer).count(),
        "policies": db.query(Policy).count(),
        "claims": db.query(Claim).count(),
        "renewals": db.query(Renewal).count(),
        "notifications": db.query(Notification).count(),
    }
